import os
import requests
import json
import pandas as pd
import tkinter as tk
from datetime import datetime
import matplotlib.pyplot as plt
from tkinter import Tk, filedialog, simpledialog, messagebox
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
import numpy as np
from PIL import Image, ImageDraw, ImageFont  # Importazioni necessarie per collage-v2.py

# Funzioni di collage-v2 integrate
def select_folder():
    root = Tk()
    root.withdraw()
    folder_selected = filedialog.askdirectory(title="Seleziona la cartella di partenza")
    return folder_selected

def custom_sort_by_order(file_list, order):
    """Ordina i file in base a un ordine personalizzato."""
    return [f for f in order if f in file_list]

def find_images_by_analysis(base_folder):
    """Trova automaticamente le immagini nelle sottocartelle di analisi breve, medio e lungo, ordinandole per nome."""
    analyses = ["breve", "medio", "lungo"]
    images_by_analysis = {}

    for analysis in analyses:
        analysis_folder = os.path.join(base_folder, f"analisi_{analysis}")
        if not os.path.exists(analysis_folder):
            continue

        for root, _, files in os.walk(analysis_folder):
            image_files = [f for f in files if f.lower().endswith(('.png', '.jpg', '.jpeg'))]
            images_by_analysis[analysis] = images_by_analysis.get(analysis, []) + [
                os.path.join(root, img) for img in image_files
            ]

    return images_by_analysis

def create_collage(company_name, analysis_type, images_by_report, output_path):
    """Crea un collage con i report allineati in righe e i dati normalizzati sotto ciascuno."""
    all_images = []
    for report, (originals, normalized) in images_by_report.items():
        originals_loaded = load_images(originals)
        normalized_loaded = load_images(normalized)
        if originals_loaded and normalized_loaded:
            all_images.append((report, originals_loaded, normalized_loaded))

    if not all_images:
        print(f"Nessuna immagine valida trovata per {analysis_type}.")
        return

    max_width = max(sum(img.width for img in originals) for _, originals, _ in all_images)
    total_height = sum(
        max(max(img.height for img in originals), max(img.height for img in normalized)) * 2
        for _, originals, normalized in all_images
    ) + 350  # Offset maggiore per il font grande

    collage = Image.new('RGB', (max_width, total_height), 'white')
    draw = ImageDraw.Draw(collage)

    try:
        font = ImageFont.truetype("arial.ttf", 24)  # Font ingrandito
    except:
        font = ImageFont.load_default()

    title_text = f"{company_name} - {analysis_type}"
    draw.text((10, 10), title_text, fill="black", font=font)

    y_offset = 30  # Maggiore spazio sopra il primo report
    for report, originals, normalized in all_images:
        draw.text((10, y_offset), report, fill="black", font=font)
        y_offset += 30  # Maggiore spazio per il titolo del report

        # Allinea le immagini in riga per originals
        x_offset = 0
        max_orig_height = max((img.height for img in originals), default = 0)
        for img in originals:
            collage.paste(img, (x_offset, y_offset))
            x_offset += img.width + 10
        y_offset += 200  #Maggiore Spazio immagini sotto testo
        
        # Allinea le immagini in riga per normalized
        x_offset = 0
        max_norm_height = max((img.height for img in normalized), default = 0)
        for img in normalized:
            collage.paste(img, (x_offset, y_offset))
            x_offset += img.width + 10
        y_offset += 150 #Maggiore spazio dopo le immagini
    
    collage.save(output_path)
    print(f"Collage salvato in: {output_path}")

def download_and_save_report(api_key, ticker, report_type, file_name):
    url = f"https://www.alphavantage.co/query?function={report_type}&symbol={ticker}&apikey={api_key}"
    response = requests.get(url)
    data = response.json()
    with open(file_name, 'w') as file:
        json.dump(data, file, indent=4)

def save_reports_to_folder(api_key, ticker):
    folder_path = os.path.join(os.path.expanduser("~"), "Downloads", ticker)
    os.makedirs(folder_path, exist_ok=True)

    reports = {
        "OVERVIEW": f"{ticker}_Overview.json",
        "INCOME_STATEMENT": f"{ticker}_Income_Statement.json",
        "BALANCE_SHEET": f"{ticker}_Balance_Sheet.json",
        "CASH_FLOW": f"{ticker}_Cash_Flow.json",
        "TIME_SERIES_MONTHLY_ADJUSTED": f"{ticker}_Historical_Prices_Adjusted.json",
        "EARNINGS": f"{ticker}_Earnings.json"
    }

    for report_type, file_name in reports.items():
        print(f"Downloading {report_type} report for {ticker}...")
        download_and_save_report(api_key, ticker, report_type, os.path.join(folder_path, file_name))

    print(f"All reports for '{ticker}' have been saved in the folder: {folder_path}")
    return folder_path

def json_to_excel(json_path, excel_path, sheet_name, columns_to_extract, fiscal_dates=None):
    with open(json_path) as f:
        data = json.load(f)

    if sheet_name == "OVERVIEW":
        overview_data = {
            'Symbol': data.get('Symbol'),
            'Name': data.get('Name'),
            'Description': data.get('Description'),
            'Exchange': data.get('Exchange'),
            'Country': data.get('Country'),
            'Sector': data.get('Sector'),
            'Industry': data.get('Industry'),
            'FiscalYearEnd': data.get('FiscalYearEnd'),
            'LatestQuarter': data.get('LatestQuarter'),
            'MarketCapitalization': data.get('MarketCapitalization'),
            'AnalystTargetPrice': data.get('AnalystTargetPrice'),
            'TrailingPE': data.get('TrailingPE'),
            'ForwardPE': data.get('ForwardPE'),
            'PriceToSalesRatioTTM': data.get('PriceToSalesRatioTTM'),
            'PriceToBookRatio': data.get('PriceToBookRatio'),
            'EVToRevenue': data.get('EVToRevenue'),
            'EVToEBITDA': data.get('EVToEBITDA'),
            'Beta': data.get('Beta'),
            '52WeekHigh': data.get('52WeekHigh'),
            '52WeekLow': data.get('52WeekLow'),
        }

        df = pd.DataFrame(list(overview_data.items()), columns=['Item', 'Value'])
    elif sheet_name == "TIME_SERIES_MONTHLY_ADJUSTED":
        monthly_data = data['Monthly Adjusted Time Series']
        filtered_data = []
        for date, values in monthly_data.items():
            close = float(values.get('4. close'))
            adjusted_close = float(values.get('5. adjusted close'))
            adjustment_factor = adjusted_close / close if close != 0 else 1.0
            open_adjusted = float(values.get('1. open')) * adjustment_factor
            high_adjusted = float(values.get('2. high')) * adjustment_factor
            low_adjusted = float(values.get('3. low')) * adjustment_factor
            row = {
                'Date': date,
                'Open': values.get('1. open'),
                'High': values.get('2. high'),
                'Low': values.get('3. low'),
                'Close': values.get('4. close'),
                'Adjusted Close': values.get('5. adjusted close'),
                'OpenAdjusted': open_adjusted,
                'HighAdjusted': high_adjusted,
                'LowAdjusted': low_adjusted,
                'Volume': float(values.get('6. volume'))
            }
            filtered_data.append(row)

        df = pd.DataFrame(filtered_data)
        df['Date'] = pd.to_datetime(df['Date'])
        df = df.sort_values(by='Date', ascending=True)
        df['QuarterEnd'] = df['Date'] + pd.offsets.QuarterEnd(0)  # Aggrega i dati per QuarterEnd
        quarterly_df = df.groupby('QuarterEnd').agg({
            'Open': 'first',
            'High': 'max',
            'Low': 'min',
            'Close': 'last',
            'Adjusted Close': 'last',
            'OpenAdjusted': 'first',
            'HighAdjusted': 'max',
            'LowAdjusted': 'min',
            'Volume': 'sum'
        }).reset_index()
        quarterly_df = quarterly_df.rename(columns={'QuarterEnd': 'Date'})
        quarterly_df['Date'] = quarterly_df['Date'].dt.strftime('%Y-%m-%d')
        quarterly_df = quarterly_df.sort_values(by='Date', ascending=True)
        df = quarterly_df[columns_to_extract]
    elif sheet_name == "EARNINGS":
        if 'quarterlyEarnings' in data:
            reports = data['quarterlyEarnings']
            filtered_data = []
            for report in reports:
                row = {col: report.get(col) for col in columns_to_extract}
                filtered_data.append(row)
            df = pd.DataFrame(filtered_data)
            if 'fiscalDateEnding' in df.columns:
                df['fiscalDateEnding'] = pd.to_datetime(df['fiscalDateEnding'])
                df = df.sort_values(by='fiscalDateEnding')
                df['fiscalDateEnding'] = df['fiscalDateEnding'].dt.strftime('%Y-%m-%d')
            if 'surprisePercentage' in df.columns:
                df['surprisePercentage'] = pd.to_numeric(df['surprisePercentage'], errors='coerce')
                df['surprisePercentage'] = df['surprisePercentage'].round(2)
        else:
            messagebox.showerror("Errore", f"La chiave 'quarterlyEarnings' non è presente nel file JSON: {json_path}")
            return None
    else:
        if 'quarterlyReports' in data:
            reports = data['quarterlyReports']
            filtered_data = []
            for report in reports:
                row = {col: report.get(col) for col in columns_to_extract}
                filtered_data.append(row)
            df = pd.DataFrame(filtered_data)
            if 'fiscalDateEnding' in df.columns:
                df = df.sort_values(by='fiscalDateEnding', ascending=True)
        else:
            messagebox.showerror("Errore", f"La chiave 'quarterlyReports' non è presente nel file JSON: {json_path}")
            return None

    if sheet_name != "OVERVIEW":
        for col in df.columns:
            if col != 'Date' and col != 'fiscalDateEnding':
                df[col] = pd.to_numeric(df[col], errors='coerce')

    try:
        df.to_excel(excel_path, sheet_name=sheet_name, index=False)
    except PermissionError:
        messagebox.showerror("Errore", f"Impossibile scrivere il file {excel_path}. Assicurati che il file non sia aperto in un altro programma.")
        return None
    except Exception as e:
        messagebox.showerror("Errore", f"Si è verificato un errore durante la scrittura del file Excel: {e}")
        return None

    return df

def format_excel(file_path, sheet_name):
    try:
        wb = load_workbook(file_path)
        ws = wb[sheet_name]

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter

            for cell in col:
                if sheet_name == "OVERVIEW":
                    cell.alignment = Alignment(horizontal="left")
                    if isinstance(cell.value, (int, float)):
                        if 'Price' in col[0].value:
                            cell.number_format = '"$"#,##0.00'
                        else:
                            cell.number_format = "#,##0.00"
                else:
                    cell.alignment = Alignment(horizontal="right")
                    if isinstance(cell.value, (int, float)):
                        if "Variazione" in str(col[0].value):
                            cell.number_format = '0.00'
                        else:
                            cell.number_format = '"$"#,##0.00'

                max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(file_path)
    except PermissionError:
        messagebox.showerror("Errore", f"Impossibile salvare il file {file_path}. Assicurati che il file non sia aperto in un altro programma.")
    except Exception as e:
        messagebox.showerror("Errore", f"Si è verificato un errore durante la formattazione del file Excel: {e}")

def merge_excel_files(file_paths, output_path):
    try:
        output_wb = Workbook()
        output_wb.remove(output_wb.active)

        for file_path in file_paths:
            wb = load_workbook(file_path)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                new_sheet = output_wb.create_sheet(title=sheet_name)

                for row in ws.iter_rows():
                    for cell in row:
                        new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                        if cell.has_style:
                            new_cell.font = cell.font.copy() if cell.font else None
                            new_cell.border = cell.border.copy() if cell.border else None
                            new_cell.fill = cell.fill.copy() if cell.fill else None
                            new_cell.number_format = cell.number_format
                            new_cell.protection = cell.protection.copy() if cell.protection else None
                            new_cell.alignment = cell.alignment.copy() if cell.alignment else None

        output_wb.save(output_path)
        messagebox.showinfo("Successo", f"File uniti con successo in: {output_path}")
    except PermissionError:
        messagebox.showerror("Errore", f"Impossibile salvare il file {output_path}. Assicurati che il file non sia aperto in un altro programma.")
    except Exception as e:
        messagebox.showerror("Errore", f"Si è verificato un errore durante l'unione dei file Excel: {e}")

def create_plots(df, columns, time_column, output_dir, normalize=False):
    skipped_columns = []

    # Ordinamento e formattazione delle date
    df = df.sort_values(by=time_column)
    df[time_column] = pd.to_datetime(df[time_column]).dt.strftime('%Y-%m')

    for column in columns:
        if column not in df.columns:
            continue

        plt.figure(figsize=(14, 7))
        df_sorted = df.sort_values(by=time_column, ascending=True).copy()

        if normalize:
            try:
                base_value = df_sorted[column].iloc[0]
                if pd.isna(base_value) or base_value == 0:
                    skipped_columns.append(column)
                    plt.close()
                    continue  # Normalizzazione dei dati

                try:
                    y_values = (df_sorted[column] / base_value - 1) * 100
                    y_label = f'Variazione % rispetto a {df_sorted[time_column].iloc[0]}'
                    reference_line = 0
                except Exception:
                    skipped_columns.append(column)
                    plt.close()
                    continue

                file_suffix = '_normalized'
            except Exception:
                skipped_columns.append(column)
                plt.close()
                continue
        else:
            y_values = df_sorted[column]
            y_label = column
            reference_line = None
            file_suffix = ''

        bars = plt.bar(
            df_sorted[time_column],
            y_values,
            color='#3498db' if not normalize else '#2ecc71',
            edgecolor='#2980b9' if not normalize else '#27ae60',
            alpha=0.9,
            width=0.7
        )

        for bar in bars:
            height = bar.get_height()
            display_value = f'{height:.1f}%' if normalize and 'volume' not in column.lower() else f'{height:,.1f}'
            plt.text(bar.get_x() + bar.get_width()/2., height, display_value, ha='center', va='bottom', fontsize=8)

        plt.xlabel('Data', fontsize=12)
        plt.ylabel(y_label, fontsize=12)
        plt.title(f'{column} nel tempo{" (Normalizzato)" if normalize else ""}', fontsize=14)

        if reference_line is not None:
            plt.axhline(reference_line, color='gray', linestyle='--', linewidth=1)

        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        plt.savefig(os.path.join(output_dir, f'{column}{file_suffix}.png'), dpi=150)
        plt.close()

    return skipped_columns

def analyze_data(folder_path, years):
    data_definitions = {
        "INCOME_STATEMENT": ["totalRevenue", "netIncome", "ebitda", "operatingExpenses", "researchAndDevelopment"],
        "BALANCE_SHEET": ["totalAssets", "totalCurrentAssets", "inventory", "totalLiabilities", "totalCurrentLiabilities", "retainedEarnings"],
        "CASH_FLOW": ["operatingCashflow", "capitalExpenditures", "cashflowFromInvestment", "cashflowFromFinancing", "netIncome"],
        "TIME_SERIES_MONTHLY_ADJUSTED": ["LowAdjusted", "HighAdjusted", "Adjusted Close", "Volume"],
        "EARNINGS": ["reportedEPS", "estimatedEPS", "surprise", "surprisePercentage"]
    }

    base_dir = folder_path
    output_dirs = {period: os.path.join(base_dir, f'analisi_{period}') for period in ["breve", "medio", "lungo"]}
    for dir_path in output_dirs.values():
        os.makedirs(dir_path, exist_ok=True)

    all_skipped = []

    for sheet_name, columns in data_definitions.items():
        file_path = os.path.join(folder_path, f"{ticker}_{sheet_name}.xlsx")
        if not os.path.exists(file_path):
            continue

        df = pd.read_excel(file_path)
        if sheet_name == "TIME_SERIES_MONTHLY_ADJUSTED":
            time_column = "Date"
        else:
            time_column = "fiscalDateEnding"

        df[time_column] = pd.to_datetime(df[time_column], errors='coerce')
        df = df.dropna(subset=[time_column])

        for period, year in years.items():
            filtered_df = df[df[time_column].dt.year >= year]
            if filtered_df.empty:
                continue

            output_dir_original = os.path.join(output_dirs[period], sheet_name, 'originale')
            output_dir_normalized = os.path.join(output_dirs[period], sheet_name, 'normalizzato')
            os.makedirs(output_dir_original, exist_ok=True)
            os.makedirs(output_dir_normalized, exist_ok=True)

            skipped = create_plots(filtered_df, columns, time_column, output_dir_normalized, normalize=True)
            create_plots(filtered_df, columns, time_column, output_dir_original)
            all_skipped.extend(skipped)

    if all_skipped:
        messagebox.showwarning("Avviso", f"Alcune colonne non sono state normalizzate:\n{', '.join(set(all_skipped))}")
    # messagebox.showinfo("Completato", "Elaborazione terminata!") #rimosso perchè il messaggio viene visualizzato prima del collage
    
def load_images(image_paths):
    """Carica le immagini da una lista di percorsi."""
    loaded_images = []
    for path in image_paths:
        try:
            img = Image.open(path)
            loaded_images.append(img)
        except FileNotFoundError:
            print(f"File non trovato: {path}")
        except Exception as e:
            print(f"Errore durante il caricamento dell'immagine {path}: {e}")
    return loaded_images

def main():
    root = Tk()
    root.withdraw()

    # GUI per API Key, Ticker e Anni
    api_key = simpledialog.askstring("Alpha Vantage API Key", "Inserisci la tua API Key di Alpha Vantage:")
    if not api_key:
        messagebox.showinfo("Operazione annullata", "Nessuna API Key inserita.")
        return

    global ticker
    ticker = simpledialog.askstring("Ticker", "Inserisci il ticker della compagnia:").upper()
    if not ticker:
        messagebox.showinfo("Operazione annullata", "Nessun ticker inserito.")
        return

    years = {}
    for period in ["breve", "medio", "lungo"]:
        year = simpledialog.askinteger(f"Inserisci anno di riferimento {period} termine", f"Anno di inizio ({period} termine):")
        if year is None:
            messagebox.showinfo("Operazione annullata", "Anni non inseriti.")
            return
        years[period] = year

    # Scarica e elabora i dati
    try:
        folder_path = save_reports_to_folder(api_key, ticker)

        json_files = {
            "OVERVIEW": ["Symbol", "Name", "Description", "Exchange", "Country", "Sector", "Industry", "FiscalYearEnd", "LatestQuarter", "MarketCapitalization", "AnalystTargetPrice", "TrailingPE", "ForwardPE", "PriceToSalesRatioTTM", "PriceToBookRatio", "EVToRevenue", "EVToEBITDA", "Beta", "52WeekHigh", "52WeekLow"],
            "INCOME_STATEMENT": ["fiscalDateEnding", "totalRevenue", "netIncome", "ebitda", "operatingExpenses", "researchAndDevelopment"],
            "BALANCE_SHEET": ["fiscalDateEnding", "totalAssets", "totalCurrentAssets", "inventory", "totalLiabilities", "totalCurrentLiabilities", "retainedEarnings", "commonStockSharesOutstanding"],
            "CASH_FLOW": ["fiscalDateEnding", "operatingCashflow", "capitalExpenditures", "cashflowFromInvestment", "cashflowFromFinancing", "netIncome"],
            "TIME_SERIES_MONTHLY_ADJUSTED": ["Date", "Open", "High", "Low", "Close", "OpenAdjusted", "LowAdjusted", "HighAdjusted", "Adjusted Close", "Volume"],
            "EARNINGS": ["fiscalDateEnding", "reportedDate", "reportedEPS", "estimatedEPS", "surprise", "surprisePercentage", "reportTime"]
        }

        excel_files = []

        # Process reports to collect fiscal dates
        for report_type in ["OVERVIEW", "INCOME_STATEMENT", "BALANCE_SHEET", "CASH_FLOW", "TIME_SERIES_MONTHLY_ADJUSTED", "EARNINGS"]:
            columns = json_files[report_type]
            json_file_name = f"{ticker}_{report_type}.json"
            if report_type == "TIME_SERIES_MONTHLY_ADJUSTED":
                json_file_name = f"{ticker}_Historical_Prices_Adjusted.json"
            json_path = os.path.join(folder_path, json_file_name)
            excel_file_name = f"{ticker}_{report_type}.xlsx"
            excel_path = os.path.join(folder_path, excel_file_name)
            df = json_to_excel(json_path, excel_path, report_type, columns)
            if df is not None:
                format_excel(excel_path, report_type)
                excel_files.append(excel_path)

        # Merge Excel files
        merged_file_path = os.path.join(folder_path, f"{ticker}_Merged.xlsx")
        merge_excel_files(excel_files, merged_file_path)

        # Avvia l'analisi dei dati (report4)
        analyze_data(folder_path, years)

        messagebox.showinfo("Completato", "Scaricamento, elaborazione e creazione file merged completati!")

        # Definisci l'ordine dei report
        report_order = ["TIME_SERIES_MONTHLY_ADJUSTED", "INCOME_STATEMENT", "BALANCE_SHEET", "CASH_FLOW", "EARNINGS"]

        #Definizione dell'ordine delle immagini per ogni report
        image_order_ORIGINALI = {
            "TIME_SERIES_MONTHLY_ADJUSTED": ["Adjusted Close.png", "Volume.png", "LowAdjusted.png", "HighAdjusted.png"],
            "INCOME_STATEMENT": ["totalRevenue.png", "netIncome.png", "ebitda.png", "operatingExpenses.png", "researchAndDevelopment.png"],
            "BALANCE_SHEET": ["totalAssets.png", "totalLiabilities.png", "totalCurrentAssets.png", "totalCurrentLiabilities.png", "retainedEarnings.png", "inventory.png"],
            "CASH_FLOW": ["operatingCashflow.png", "cashflowFromFinancing.png", "cashflowFromInvestment.png", "capitalExpenditures.png"],
            "EARNINGS": ["reportedEPS.png", "estimatedEPS.png", "surprise.png", "surprisePercentage.png"]
        }

        image_order_NORMALIZZATE = {
            "TIME_SERIES_MONTHLY_ADJUSTED": ["Adjusted Close_normalized.png", "Volume_normalized.png", "LowAdjusted_normalized.png", "HighAdjusted_normalized.png"],
            "INCOME_STATEMENT": ["totalRevenue_normalized.png", "netIncome_normalized.png", "ebitda_normalized.png", "operatingExpenses_normalized.png", "researchAndDevelopment_normalized.png"],
            "BALANCE_SHEET": ["totalAssets_normalized.png", "totalLiabilities_normalized.png", "totalCurrentAssets_normalized.png", "totalCurrentLiabilities_normalized.png", "retainedEarnings_normalized.png", "inventory_normalized.png"],
            "CASH_FLOW": ["operatingCashflow_normalized.png", "cashflowFromFinancing_normalized.png", "cashflowFromInvestment_normalized.png", "capitalExpenditures_normalized.png"],
            "EARNINGS": ["reportedEPS_normalized.png", "estimatedEPS_normalized.png", "surprise_normalized.png", "surprisePercentage_normalized.png"]
        }
        def load_images(image_paths):
            """Carica le immagini da una lista di percorsi."""
            loaded_images = []
            for path in image_paths:
                try:
                    img = Image.open(path)
                    loaded_images.append(img)
                except FileNotFoundError:
                    print(f"File non trovato: {path}")
                except Exception as e:
                    print(f"Errore durante il caricamento dell'immagine {path}: {e}")
            return loaded_images

        def create_collage(company_name, analysis_type, images_by_report, output_path):
            """Crea un collage con i report allineati in righe e i dati normalizzati sotto ciascuno."""

            all_images = []
            for report, (originals, normalized) in images_by_report.items():
                originals_loaded = load_images([os.path.join(base_folder, "analisi_" + analysis_type, report, "originale", original) for original in originals if os.path.exists(os.path.join(base_folder, "analisi_" + analysis_type, report, "originale", original))])
                normalized_loaded = load_images([os.path.join(base_folder, "analisi_" + analysis_type, report, "normalizzato", normalized_img) for normalized_img in normalized if os.path.exists(os.path.join(base_folder, "analisi_" + analysis_type, report, "normalizzato", normalized_img))])
                if originals_loaded and normalized_loaded:
                    all_images.append((report, originals_loaded, normalized_loaded))

            if not all_images:
                print(f"Nessuna immagine valida trovata per {analysis_type}.")
                return

            max_width = max(sum(img.width for img in originals) for _, originals, _ in all_images)
            total_height = sum(
                max(max(img.height for img in originals), max(img.height for img in normalized)) * 2
                    for _, originals, normalized in all_images
            ) + 1500  # Offset maggiore per il font grande
    
            collage = Image.new('RGB', (max_width, total_height), 'white')
            draw = ImageDraw.Draw(collage)

            try:
                font = ImageFont.truetype("arial.ttf", 24)  # Font ingrandito
            except:
                font = ImageFont.load_default()

            title_text = f"{company_name} - {analysis_type}"
            draw.text((10, 10), title_text, fill="black", font=font)

            y_offset = 350  # Maggiore spazio sopra il primo report
            for report, originals, normalized in all_images:
                draw.text((10, y_offset), report, fill="black", font=font)
                y_offset += 200  # Maggiore spazio per il titolo del report
        
                # Allinea le immagini in riga per originals
                x_offset = 0
                max_orig_height = max(img.height for img in originals)
                for img in originals:
                    collage.paste(img, (x_offset, y_offset))
                    x_offset += img.width
                y_offset += max_orig_height
                
                # Allinea le immagini in riga per normalized
                x_offset = 0
                max_norm_height = max(img.height for img in normalized)
                for img in normalized:
                    collage.paste(img, (x_offset, y_offset))
                    x_offset += img.width
                y_offset += max_norm_height
                          
            collage.save(output_path)
            print(f"Collage salvato in: {output_path}")

        # Integrazione del programma di collage: avvia la creazione del collage utilizzando i dati generati
        messagebox.showinfo("Completato", "Elaborazione terminata!")

        base_folder = folder_path  # Cartella base contenente le cartelle di analisi
        output_folder = base_folder  # Cartella dove salvare il collage

        for analysis_type in ["breve", "medio", "lungo"]:
            images_by_report = {}
            for report in report_order:
                images_by_report[report] = (image_order_ORIGINALI.get(report, []), image_order_NORMALIZZATE.get(report, []))

            output_collage_path = os.path.join(output_folder, f"collage_{analysis_type}.jpg")
            create_collage(ticker, analysis_type, images_by_report, output_collage_path)

        messagebox.showinfo("Collage Completato", "I collage sono stati creati con successo!")

    except Exception as e:
        messagebox.showerror("Errore", str(e))
        
    def percent_change(old, new):
        return ((new - old) / abs(old)) * 100 if old != 0 else None  
    file_path = os.path.join(folder_path, f"{ticker}_Merged.xlsx")
    print(f"File path: {file_path}") # Debugging
    # Check if the file exists
    if not os.path.exists(file_path):
        print(f"Error: File not found at {file_path}")
        exit()
    try:
        num_years = 15
        if num_years <= 0:
            raise ValueError
    except ValueError:
        print("Numero di anni non valido. Uscita.")
        exit()
    xls = pd.ExcelFile(file_path)
    sheets = [s for s in xls.sheet_names if s not in {"OVERVIEW", "TIME_SERIES_MONTHLY_ADJUSTED"}]
    # Caricamento dati quotazione
    quotes = pd.read_excel(xls, sheet_name="TIME_SERIES_MONTHLY_ADJUSTED")
    quotes.columns = quotes.columns.str.strip()
    quotes["Date"] = pd.to_datetime(quotes["Date"], errors='coerce').dt.normalize()
    quotes = quotes.dropna(subset=["Date"]).sort_values("Date")
    quotes["Year"] = quotes["Date"].dt.year
    if quotes.empty:
        print("Nessun dato disponibile nella quotazione.")
        exit()
    current_year = datetime.now().year
    available_years = sorted(quotes["Year"].unique(), reverse=True)[:num_years]
    # Preparazione struttura dati
    all_data = []
    stock_prices = [] 
    # Elaborazione quotazioni
    for year in available_years:
        year_data = quotes[quotes["Year"] == year]
        if not year_data.empty:
            start = year_data.iloc[0]["Adjusted Close"]
            latest = quotes.iloc[-1]["Adjusted Close"]
            change = percent_change(start, latest)
            stock_prices.append({
                "Report": "Stock Price (Adjusted Close)",
                "Year": year,
                "Start Value": start,
                "Latest Value": latest,
                "% Change": change
            })
    # Elaborazione report
    for sheet in sheets:
        try:
            df = pd.read_excel(xls, sheet_name=sheet)
            df.columns = df.columns.str.strip()
            if 'fiscalDateEnding' not in df.columns:
                continue  
            df["fiscalDateEnding"] = pd.to_datetime(df["fiscalDateEnding"], errors='coerce').dt.normalize()
            df = df.dropna(subset=["fiscalDateEnding"]).sort_values("fiscalDateEnding")
            df["Year"] = df["fiscalDateEnding"].dt.year
            latest_values = df.iloc[-1]
            for year in available_years:
                year_data = df[df["Year"] == year]
                if not year_data.empty:
                    start_values = year_data.iloc[0]
                    for col in df.columns:
                        if col not in {"fiscalDateEnding", "Year"}:
                            all_data.append({
                                "Report": f"{sheet} - {col}",
                                "Year": year,
                                "Start Value": start_values[col],
                                "Latest Value": latest_values[col],
                                "% Change": percent_change(start_values[col], latest_values[col])
                            })
        except Exception as e:
            print(f"Errore nel processare {sheet}: {str(e)}") 
    # Creazione DataFrame
    combined_df = pd.DataFrame(stock_prices + all_data)
    # Gestione anni mancanti
    all_years = list(range(current_year, current_year - num_years, -1))
    multi_index = pd.MultiIndex.from_product(
        [combined_df['Report'].unique(), all_years],
        names=['Report', 'Year']
    )
    combined_df = combined_df.set_index(['Report', 'Year']).reindex(multi_index).reset_index()
    # Ordina i report
    report_order = ["Stock Price (Adjusted Close)"] + sorted(
        [r for r in combined_df['Report'].unique() if r != "Stock Price (Adjusted Close)"]
    )
    combined_df['Report'] = pd.Categorical(combined_df['Report'], categories=report_order, ordered=True)
    combined_df = combined_df.sort_values(['Year', 'Report'], ascending=[False, True])
    # Salvataggio con formattazione
    
    base_folder = folder_path  # Cartella base contenente le cartelle di analisi
    output_folder = base_folder  # Cartella dove salvare il collage
    output_file = os.path.join(output_folder, f"{ticker}_Financial_Report.xlsx")
 
    with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
        workbook = writer.book
        worksheet = workbook.add_worksheet('Report')  
        # Formattazione
        header_format = workbook.add_format({
            'bold': True, 
            'align': 'center', 
            'valign': 'vcenter',
            'border': 1
        })
        currency_format = workbook.add_format({'num_format': '$#,##0.00'})
        percent_format = workbook.add_format({'num_format': '0.00%'})
        bold_format = workbook.add_format({'bold': True}) 
        row_idx = 0
        for year in sorted(combined_df['Year'].unique(), reverse=True):
            # Intestazione anno
            worksheet.merge_range(row_idx, 0, row_idx, 3, f'FROM {year}', header_format)
            row_idx += 1  
            # Intestazioni colonne
            worksheet.write(row_idx, 0, "Report", bold_format)
            worksheet.write(row_idx, 1, "Start Value", bold_format)
            worksheet.write(row_idx, 2, "Latest Value", bold_format)
            worksheet.write(row_idx, 3, "% Change", bold_format)
            row_idx += 1  
            # Dati
            year_data = combined_df[combined_df['Year'] == year]
            for _, row in year_data.iterrows():
                # Valori
                start_val = row['Start Value'] if not pd.isna(row['Start Value']) else 'N/D'
                latest_val = row['Latest Value'] if not pd.isna(row['Latest Value']) else 'N/D'
                change_val = row['% Change'] if not pd.isna(row['% Change']) else 'N/D'
                # Scrittura
                worksheet.write(row_idx, 0, row['Report'])
                if isinstance(start_val, (int, float)):
                    worksheet.write(row_idx, 1, start_val, currency_format)
                else:
                    worksheet.write(row_idx, 1, start_val)
                if isinstance(latest_val, (int, float)):
                    worksheet.write(row_idx, 2, latest_val, currency_format)
                else:
                    worksheet.write(row_idx, 2, latest_val)
                if isinstance(change_val, (int, float)):
                    worksheet.write(row_idx, 3, change_val/100, percent_format)
                else:
                    worksheet.write(row_idx, 3, change_val)
                row_idx += 1  
            # Spaziatura tra anni
            row_idx += 1  
        # Regolazione larghezza colonne
        worksheet.set_column(0, 0, 45)   # Report
        worksheet.set_column(1, 2, 18)   # Valori
        worksheet.set_column(3, 3, 12)   # %  
    print(f"Report generato: {output_file}")

if __name__ == "__main__":
    main()
